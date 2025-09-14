import re
import json
import requests
import random
from bs4 import BeautifulSoup
import os
import urllib.parse
import openpyxl
from openpyxl.styles import PatternFill
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials as UserCredentials
from openai import OpenAI

# ================================
# OpenAI API 키 로드
# ================================
OPENAI_API_KEY = ""
if os.path.exists("openai.json"):
    with open("openai.json", "r", encoding="utf-8") as f:
        data = json.load(f)
        OPENAI_API_KEY = data.get("api_key", "").strip()
if not OPENAI_API_KEY:
    OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()

client = OpenAI(api_key=OPENAI_API_KEY)

# ================================
# 1. 엑셀에서 URL 추출
# ================================
filename = "복지서비스목록.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.active

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

target_row = None
my_url = None

for row in ws.iter_rows(min_row=2):  # 2행부터 (헤더 제외)
    url_cell = row[4]    # F열 (상세URL)
    status_cell = row[7] # H열 ("완" 여부)

    if url_cell.value and (not status_cell.value or status_cell.value.strip() != "완"):
        my_url = url_cell.value
        target_row = row
        break

if not my_url:
    print("🔔 처리할 새로운 URL이 없습니다. (모든 행 완료됨)")
    exit()

print("👉 이번에 처리할 URL:", my_url)

# 쿼리스트링 파싱
parsed = urllib.parse.urlparse(my_url)
params = urllib.parse.parse_qs(parsed.query)
wlfareInfoId = params.get("wlfareInfoId", [""])[0]
print("wlfareInfoId =", wlfareInfoId)

# ================================
# 2. 복지 서비스 데이터 가져오기
# ================================
def fetch_welfare_info(wlfareInfoId):
    url = f"https://www.bokjiro.go.kr/ssis-tbu/twataa/wlfareInfo/moveTWAT52011M.do?wlfareInfoId={wlfareInfoId}&wlfareInfoReldBztpCd=01"
    resp = requests.get(url)
    resp.encoding = "utf-8"
    html = resp.text
    outer_match = re.search(r'initParameter\((\{.*?\})\);', html, re.S)
    if not outer_match:
        raise ValueError("initParameter JSON을 찾지 못했습니다.")
    outer_data = json.loads(outer_match.group(1))
    inner_str = outer_data["initValue"]["dmWlfareInfo"]
    return json.loads(inner_str)

def clean_html(raw_html):
    return BeautifulSoup(raw_html, "html.parser").get_text(separator="\n", strip=True)

# ================================
# 3. ChatGPT API로 본문 가공
# ================================
def process_with_gpt(section_title: str, raw_text: str, keyword: str, icon: str = "📌") -> str:
    system_msg = (
        "너는 한국어 블로그 글을 쓰는 카피라이터야. "
        "출력 형식은 반드시 아래 구조로 만들어야 한다:\n\n"
        "<section class=\"custom-section\">\n"
        "<h2 class=\"section-title\">[아이콘] [섹션 제목]</h2>\n"
        "[본문 내용]\n"
        "<p><br /></p><p><br /></p>\n"
        "</section>\n\n"
        "본문 작성 규칙:\n"
        "1) 첫 <p>에는 <b>태그로 한두 문장 요약</b>\n"
        "2) 이어지는 <p>들은 친절하고 상세한 설명\n"
        "3) 필요 시 <strong>강조</strong>, 줄바꿈 <br /> 사용\n"
        "4) 마크다운 절대 금지, HTML만 사용\n"
        "5) 링크는 <a href=\"URL\" target=\"_blank\">텍스트</a> 형식"
    )
    user_msg = f"[섹션 제목] {keyword} {section_title}\n[원문]\n{raw_text}"

    resp = client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.7,
        max_tokens=900,
    )
    return resp.choices[0].message.content.strip()

# ================================
# 4. Blogger 인증 (refresh_token JSON 방식)
# ================================
def get_blogger_service():
    creds = UserCredentials.from_authorized_user_file(
        "blogger_token.json",
        ["https://www.googleapis.com/auth/blogger"]
    )
    if not creds or not creds.valid:
        creds.refresh(Request())
        with open("blogger_token.json", "w", encoding="utf-8") as f:
            f.write(creds.to_json())
    return build("blogger", "v3", credentials=creds)

# ================================
# 5. 본문 생성
# ================================
data = fetch_welfare_info(wlfareInfoId)
keyword = clean_html(data.get("wlfareInfoNm", "복지 서비스"))

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|.]', '_', name)

safe_keyword = sanitize_filename(keyword)
title = f"2025 {keyword} 지원 대상 신청방법 총정리"
mytag = " ".join([f"#{word}" for word in title.split()])
print("자동 생성된 태그:", mytag)

fields = {
    "개요": "wlfareInfoOutlCn",
    "지원대상": "wlfareSprtTrgtCn",
    "서비스내용": "wlfareSprtBnftCn",
    "신청방법": "aplyMtdDc",
    "추가정보": "etct"
}

blog_handler = get_blogger_service()

html = "<div class=\"my-unique-wrapper\">\n"

icons = {
    "개요": "🚗",
    "지원대상": "👥",
    "서비스내용": "📋",
    "신청방법": "📝",
    "추가정보": "ℹ️"
}

for title_k, key in fields.items():
    value = data.get(key, "")
    if not value or value.strip() in ["", "정보 없음"]:
        continue
    text = clean_html(value)
    icon = icons.get(title_k, "📌")
    processed_text = process_with_gpt(title_k, text, keyword, icon)
    html += processed_text

html += f"""
<div class="custom-button">
  <a href="{my_url}" target="_blank">👉 {title} 자세히 알아보기</a>
</div>
</div>
"""

# ================================
# 6. 블로그 업로드
# ================================
BLOG_ID = os.getenv("BLOG_ID", "4737456424227083027")

data_post = {
    'content': html,
    'title': title,
    'labels': ["복지", "정부지원", "복지서비스"],
    'blog': {'id': BLOG_ID},
}
posts = blog_handler.posts()
res = posts.insert(blogId=BLOG_ID, body=data_post, isDraft=False, fetchImages=True).execute()

# ================================
# 7. ✅ 엑셀 업데이트
# ================================
for cell in target_row[:6]:
    cell.fill = green_fill
target_row[7].value = "완"

wb.save(filename)
print("✅ 엑셀 표시 완료: 해당 행 A~F 녹색 + H열 '완' 기록")

print(f"[완료] 블로그 포스팅: {res['url']}")
print(title)
