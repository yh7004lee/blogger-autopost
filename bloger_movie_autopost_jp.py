#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel(MOVIE_ID) → TMDB → Blogger 自動ポスティングパイプライン
- movies_discover.xlsx 読み込み: A=タイトル, B=MOVIE_ID, C=公開日, D=評価, E=投票数, F=備考, G=完了フラグ
- G列が「完」の行はスキップし、最初の未完了行(B列のMOVIE_ID)で投稿
- TMDB 詳細/出演者/画像/レビュー/おすすめ/予告編 収集
- ランダムイントロ(6文), セクションリード(4文), アウトロ(6文)
- Blogger APIで公開 (blogId=6854008537633494036)
- 成功時に対象行G列へ「完」と記録して保存
"""
import json
import urllib.parse
import os, sys, html, textwrap, requests, random, time, pickle
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# ===============================
# 📝 ポスティング設定
POST_COUNT = 1     # 投稿回数
POST_DELAY_MIN = 1  # 各投稿後の待機時間 (分単位、0なら即次へ)

# ===============================
# 🔧 環境/パス設定
EXCEL_PATH = "movies_discover.xlsx"
BLOG_ID = "6854008537633494036"       # ★ 新しい日本語ブログ ID
CLIENT_SECRET_FILE = r"cc.json" # Google OAuth クライアントシークレットJSON
BLOGGER_TOKEN_PICKLE = "blogger_token.pickle"
SCOPES = ["https://www.googleapis.com/auth/blogger"]

# ===============================
# 🈶 TMDB 設定
LANG = "ja-JP"   # ★ 日本語
CAST_COUNT = 10
STILLS_COUNT = 8
TMDB_V3_BASE = "https://api.themoviedb.org/3"
IMG_BASE = "https://image.tmdb.org/t/p"

# 🔑 TMDB 認証情報
BEARER = "eyJhbGciOiJIUzI1NiJ9.eyJhdWQiOiI1NmY0YTNiY2UwNTEyY2RjMjAxNzFhODMxNTNjMjVkNiIsIm5iZiI6MTc1NjY0NjE4OC40MTI5OTk5LCJzdWIiOiI2OGI0NGIyYzI1NzIyYjIzNDdiNGY0YzQiLCJzY29wZXMiOlsiYXBpX3JlYWQiXSwidmVyc2lvbiI6MX0.ShX_ZJwMuZ1WffeUR6PloXx2E7pjBJ4nAlQoI4l7nKY"
API_KEY = "56f4a3bce0512cdc20171a83153c25d6"

# 🔑 YouTube API 認証
YOUTUBE_API_KEY = "AIzaSyD92QjYwV12bmLdUpdJU1BpFX3Cg9RwN4o"
YOUTUBE_SEARCH_URL = "https://www.googleapis.com/youtube/v3/search"



# ===============================
# TMDB コレクター（不足分の関数を追加）

def get_movie_videos_all(movie_id, bearer=None, api_key=None):
    """언어 우선순위를 두고 예고편을 가져오기 (ja-JP → ko-KR → en-US)"""
    videos = []

    for lang in ["ja-JP", "ko-KR", "en-US"]:
        try:
            j = tmdb_get(
                f"/movie/{movie_id}/videos",
                params={"language": lang},
                bearer=bearer,
                api_key=api_key
            )
            results = j.get("results", [])
            if results:
                # Trailer / Official Trailer / Teaser 만 필터
                filtered = [
                    v for v in results
                    if v.get("site") == "YouTube" and v.get("type") in ("Trailer", "Official Trailer", "Teaser")
                ]
                if filtered:
                    videos.extend(filtered)
                    break  # 첫 번째로 찾은 언어 버전만 사용
        except Exception as e:
            print(f"❌ TMDB video fetch error ({lang}):", e)
            continue

    return videos



def get_movie_bundle(movie_id, lang="ja-JP", bearer=None, api_key=None):
    """映画の詳細・出演・画像をまとめて取得"""
    params = {
        "language": lang,
        "append_to_response": "credits,images",
        "include_image_language": "ja,en,null"
    }
    return tmdb_get(f"/movie/{movie_id}", params=params, bearer=bearer, api_key=api_key)

def get_movie_reviews(movie_id, lang="ja-JP", bearer=None, api_key=None):
    """（将来使う場合に備えて）レビュー一覧"""
    j = tmdb_get(f"/movie/{movie_id}/reviews", params={"language": lang}, bearer=bearer, api_key=api_key)
    return j.get("results", [])

def get_movie_videos(movie_id, lang="ja-JP", bearer=None, api_key=None):
    """動画（予告編/ティーザー等）一覧"""
    j = tmdb_get(f"/movie/{movie_id}/videos", params={"language": lang}, bearer=bearer, api_key=api_key)
    return j.get("results", [])

def get_movie_recommendations(movie_id, lang="ja-JP", bearer=None, api_key=None):
    """（将来使う場合に備えて）おすすめ作品一覧"""
    j = tmdb_get(f"/movie/{movie_id}/recommendations", params={"language": lang}, bearer=bearer, api_key=api_key)
    return j.get("results", [])

def get_movie_release_cert(movie_id, bearer=None, api_key=None):
    """日本のレーティング（なければUS→空文字）"""
    data = tmdb_get(f"/movie/{movie_id}/release_dates", bearer=bearer, api_key=api_key)
    results = data.get("results", []) or []

    def find_cert(country_code):
        for r in results:
            if r.get("iso_3166_1") == country_code:
                for d in r.get("release_dates", []) or []:
                    cert = (d.get("certification") or "").strip()
                    if cert:
                        return cert
        return ""

    jp = find_cert("JP")
    if jp:
        return jp
    us = find_cert("US")
    if us:
        return us
    return ""


# ===============================
# ハッシュタグ生成 (タイトルから)
def make_hashtags_from_title(title: str) -> str:
    import re
    words = re.findall(r"[ぁ-んァ-ン一-龥A-Za-z0-9]+|\([^)]+\)", title)
    hashtags = ["#" + w.strip() for w in words if w.strip()]

    # 後処理で自然な形に修正
    fixed = []
    for h in hashtags:
        if h == "#ダンサ":
            h = "#リトルダンサー"
        if h.startswith("#(") and h.endswith(")"):
            year = h.strip("#()")
            if year.isdigit():
                h = f"#{year}年公開"
        fixed.append(h)

    return " ".join(fixed)


# ===============================
# YouTube 予告編検索
def get_youtube_trailers(query, max_results=2):
    params = {
        "part": "snippet",
        "q": query,
        "key": YOUTUBE_API_KEY,
        "maxResults": max_results,
        "type": "video",
        "videoEmbeddable": "true"
    }
    try:
        r = requests.get(YOUTUBE_SEARCH_URL, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        videos = []
        for item in data.get("items", []):
            vid = item["id"]["videoId"]
            title = item["snippet"]["title"]
            videos.append((vid, title))
        return videos
    except Exception as e:
        print(f"❌ YouTube API エラー: {e}")
        return []

# ===============================
# TMDB 共通ユーティル
def tmdb_get(path, params=None, bearer=None, api_key=None):
    url = f"{TMDB_V3_BASE}{path}"
    headers = {"Accept": "application/json"}
    if bearer:
        headers["Authorization"] = f"Bearer {bearer}"
    if params is None:
        params = {}
    if api_key and "api_key" not in params and not bearer:
        params["api_key"] = api_key
    r = requests.get(url, headers=headers, params=params, timeout=20)
    r.raise_for_status()
    return r.json()

def img_url(path, size="w780"):
    if not path:
        return None
    return f"{IMG_BASE}/{size}{path}"

def choose(*options):
    return random.choice(options)

def maybe(value, prob=0.5):
    return value if random.random() < prob else ""


# ===============================
# 🎬 イントロ生成 (6文ランダム, 日本語)
def make_intro_6(title, year, genres_str, director_names, main_cast, cert_label, runtime_min, keywords):
    title_bold = f"{title}"
    year_txt = f"{year}年公開" if year else "公開年不明"
    genre_phrase = genres_str if genres_str else "ジャンル不明"
    director_one = director_names[0] if director_names else ""
    star_one = main_cast[0] if main_cast else ""
    star_two = main_cast[1] if len(main_cast) > 1 else ""
    runtime_txt = f"{runtime_min}分" if runtime_min else "上映時間不明"
    cert_txt = cert_label or "レーティング不明"

    # 1. オープニング（20+）
    s1 = choose(
        f"こんにちは！今日は {year_txt} の話題作 <b>{title_bold}</b> をご紹介します。",
        f"映画 <b>{title_bold}</b> を知っていますか？ {year_txt} に公開された感動作です。",
        f"今回はずっと語りたかった <b>{title_bold}</b> を取り上げます。",
        f"{year_txt} に公開された <b>{title_bold}</b> を一緒に振り返りましょう。",
        f"<b>{title_bold}</b> はもう観ましたか？まだならこの記事で雰囲気をつかんでくださいね。",
        f"今日は特別に <b>{title_bold}</b> の魅力をたっぷりお話しします。",
        f"映画好きの私が推したい一作、<b>{title_bold}</b>({year_txt}) をピックアップしました。",
        f"このブログでは久々に心からおすすめしたい映画、<b>{title_bold}</b> を紹介します。",
        f"最近改めて観直して感動した <b>{title_bold}</b> のレビューです。",
        f"映画ファンなら絶対外せない <b>{title_bold}</b> を取り上げます。",
        f"今日はリラックスしながら <b>{title_bold}</b> の世界を一緒に楽しみましょう。",
        f"観た人の心に強く残る <b>{title_bold}</b>。その理由をお伝えします。",
        f"数ある映画の中でも <b>{title_bold}</b> は特別な存在です。",
        f"思い出すだけで胸が熱くなる <b>{title_bold}</b> の魅力を語ります。",
        f"公開から年月が経っても色褪せない名作、<b>{title_bold}</b>。",
        f"今日は家で映画を観るならこれ！<b>{title_bold}</b> をおすすめします。",
        f"この記事では私の大好きな <b>{title_bold}</b> をわかりやすくまとめました。",
        f"長年愛され続ける映画 <b>{title_bold}</b> をご紹介します。",
        f"まだ観ていない人にぜひ届けたい作品 <b>{title_bold}</b>。",
        f"{year_txt} に公開され、多くの人を魅了した <b>{title_bold}</b> をレビューします。"
    )

    # 2. ジャンルと雰囲気（20+）
    s2 = choose(
        f"ジャンルは {genre_phrase}。気軽に観られるのに深みもあります。",
        f"{genre_phrase} 好きにはたまらない内容です。",
        f"全体の雰囲気は {choose('落ち着いている', '温かい', 'テンポが良い', '緊張感に満ちている')} 作品です。",
        f"ジャンルが {genre_phrase} なので、すぐに世界観へ引き込まれます。",
        f"{genre_phrase} らしさが詰まっていて、心に残ります。",
        f"{genre_phrase} 作品として完成度が高いです。",
        f"物語の流れが自然で {genre_phrase} らしい楽しさがあります。",
        f"{genre_phrase} 作品の中でも特に印象に残る一作です。",
        f"{genre_phrase} の醍醐味をしっかり味わえます。",
        f"観終わったあと {genre_phrase} 好きでよかったと思える作品です。",
        f"ジャンル要素が自然に混ざり、飽きません。",
        f"{genre_phrase} だけど予想外の展開も楽しめます。",
        f"{genre_phrase} の新しい魅力を発見できる映画です。",
        f"{genre_phrase} 好きに自信を持っておすすめできます。",
        f"ジャンルの枠を超えて楽しめる作品です。",
        f"{genre_phrase} 作品の中でも特に観やすいです。",
        f"ジャンル特有の演出が活きています。",
        f"{genre_phrase} らしい余韻が残る一作です。",
        f"気分転換にぴったりの {genre_phrase} 映画です。",
        f"ジャンルの魅力をぎゅっと詰め込んだ作品です。"
    )

    # 3. 監督について（15+）
    s3 = (
        choose(
            f"監督は {director_one}。{choose('独自のセンス', '細やかな演出', '安定したテンポ')} が際立ちます。",
            f"{director_one} 監督ならではの色が作品全体に反映されています。",
            f"{director_one} 監督の手腕が光っています。",
            f"{director_one} の演出スタイルが物語を魅力的にしています。",
            f"細部までこだわる {director_one} 監督の感性が伝わります。",
            f"{director_one} 監督の作品らしさが強く感じられます。",
            f"{director_one} の手掛ける映画はいつも世界観が濃厚です。",
            f"映像のリズム感は {director_one} 監督ならではです。",
            f"演出から {director_one} の情熱が伝わります。",
            f"{director_one} 監督作品が好きな方にも響くでしょう。",
            f"{director_one} の作風が好きなら絶対に楽しめます。",
            f"{director_one} の映画はいつも人間味にあふれています。",
            f"{director_one} 監督の安定した映像美が心地よいです。",
            f"{director_one} 監督の丁寧な演出に引き込まれます。",
            f"{director_one} の個性が随所に表れています。"
        ) if director_one else choose(
            "演出は全体的に自然で観やすいです。",
            "映像と物語がバランスよくまとまっています。",
            "派手すぎず落ち着いた演出が魅力です。",
            "演出がシンプルで作品に集中できます。",
            "監督の個性は控えめですが安定感があります。"
        )
    )

    # 4. 出演者について（15+）
    s4 = (
        choose(
            f"キャストも豪華で、{star_one}{('・'+star_two) if star_two else ''} などが出演しています。",
            f"特に {star_one} の演技が光っていました。" if star_one else "出演者全員が自然な演技を見せてくれます。",
            f"{star_one} と {star_two} の相性も良く、物語を引き立てます。" if star_two else f"{star_one} の存在感が印象的でした。",
            "俳優同士の掛け合いが自然で物語に厚みを与えています。",
            "演技の呼吸が合っていてキャラクターが生き生きしています。",
            "役者の表情や動きに引き込まれます。",
            "出演者それぞれの個性が輝いています。",
            "観客を引き込むリアルな演技が魅力です。",
            "演技の自然さに驚かされます。",
            "キャスト陣の熱演に胸が熱くなります。",
            "脇役までしっかり存在感があります。",
            "子役の演技も印象的です。",
            "主演の魅力が全体を引っ張っています。",
            "俳優一人ひとりのキャラが際立っています。",
            "キャスト全員が作品を支えています。"
        )
    )

    # 5. 上映時間とレーティング（15+）
    s5 = choose(
        f"上映時間は {runtime_txt} でテンポよく進みます。",
        f"{runtime_txt} ですが退屈せず楽しめます。",
        f"{runtime_txt} の長さを感じさせない構成です。",
        f"{runtime_txt} はあっという間でした。",
        f"時間配分がちょうどよく感じました。"
    ) + " " + (
        choose(
            f"レーティングは {cert_txt} なので、{choose('家族でも安心して観られる作品です。', '友達や恋人と一緒に観ても楽しめます。')}",
            f"年齢制限は {cert_txt}。{choose('幅広い世代が楽しめます。', '子どもから大人まで安心して観られます。')}"
        ) if cert_label else "年齢制限は特にないので誰でも気軽に楽しめます。"
    )

    # 6. 案内とキーワード（15+）
    s6 = choose(
        "ここからは映画の魅力を詳しく見ていきましょう。",
        "次にストーリーや演出について掘り下げます。",
        "これから作品の見どころを順にご紹介します。",
        "以下ではあらすじ、出演者、スチール写真、評価、予告編をまとめます。",
        "一緒に作品の世界をさらに楽しんでいきましょう。"
    ) + " " + choose(
        f"検索キーワード: {', '.join(keywords[:6])}",
        f"関連ワード: {', '.join(keywords[:6])}",
        f"注目ワード: {', '.join(keywords[:6])}"
    )

    return " ".join([s1, s2, s3, s4, s5, s6])



# ===============================
# 🎬 セクションリード生成 (日本語, 4文ランダム)
def make_section_lead(name, title, year, genres_str, cert_label, extras=None):
    extras = extras or {}
    year_txt = f"{year}年" if year else ""
    genre_phrase = genres_str if genres_str else "ジャンル"
    cert_txt = cert_label or "レーティング不明"
    cast_top = extras.get("cast_top", [])
    who = "・".join(cast_top[:3]) if cast_top else ""
    director_one = extras.get("director_one", "")
    runtime_min = extras.get("runtime_min", None)
    runtime_txt = f"{runtime_min}分" if runtime_min else ""

    # ===============================
    # あらすじ
    if name == "あらすじ":
        base = [
            choose(
                f"映画『{title}』({year_txt}) のあらすじを、ネタバレなしでやさしくご紹介します。",
                f"これから観る人のために、<b>{title}</b> のストーリーを分かりやすく整理しました。",
                f"物語の核心を避けつつ、<b>{title}</b> の流れをまとめてみます。",
                f"まだ観ていない方でも安心して読めるように、主要な部分だけをピックアップしました。",
                f"<b>{title}</b>({year_txt}) の世界観を短くまとめるとこんな感じです。",
                f"初めて観る人にも分かりやすいようにポイントを整理しました。"
            ),
            choose(
                f"序盤では世界観が設定され、中盤で {choose('緊張感が高まり', '人間関係が複雑化し', '謎が深まっていき')}、",
                f"物語は {choose('出発→試練→成長', '出会い→葛藤→解決', '発端→衝突→収束')} という流れで進行します。",
                f"序盤から終盤まで流れがスムーズで分かりやすいです。"
            ),
            choose(
                f"{genre_phrase} らしい雰囲気が作品全体を彩り、トーンは {choose('落ち着いていて過剰でなく', '緊張感がありながらバランスよく', '温かさと余韻を残す')} 仕上がりです。",
                f"映像と演出の調和により、物語が自然に展開していきます。",
                f"映像と音楽のバランスが作品全体を心地よく包んでいます。"
            ),
            choose(
                f"レーティングは {cert_txt} で、{choose('家族でも安心して観られます。', '友人やカップルで楽しむのにも最適です。')}"
                if cert_label else "誰でも安心して楽しめる作品です。",
                f"{cert_txt} のため幅広い年齢層におすすめできます。"
            )
        ]

    # ===============================
    # 出演者
    elif name == "出演者":
        base = [
            choose(
                f"今回のキャストは {who} など、とても豪華です。" if who else "出演者の顔ぶれもバラエティ豊かです。",
                f"スクリーンに登場するだけで期待感が高まる俳優が揃っています。",
                f"キャストの組み合わせが絶妙で、物語をしっかり支えています。",
                f"役者陣の存在感が作品全体をより一層引き立てています。"
            ),
            choose(
                "演技の呼吸が合っていて、キャラクターが生き生きと描かれています。",
                "役者同士の掛け合いが自然で物語を引き立てます。",
                "主役も脇役も存在感が際立っています。",
                "それぞれの役がしっかりと物語に溶け込んでいます。"
            ),
            choose(
                "個性のぶつかり合いが見どころで、シーンごとに魅力が光ります。",
                "短い登場シーンでも印象的な演技を残す俳優が多いです。",
                "全体として安定感のある演技で没入感を高めています。",
                "小さな役でも観客の印象に残る演技が光っています。"
            ),
            choose(
                "ここからは主要キャストと役どころをご紹介します。",
                "それぞれの出演者がどんなキャラクターを演じたのか見ていきましょう。",
                "以下に出演者と役柄をまとめました。",
                "キャラクターごとの魅力を簡単に整理しました。"
            )
        ]

    # ===============================
    # スチール写真
    elif name == "スチール写真":
        base = [
            choose(
                "スチール写真を見るだけで映画の世界観が伝わってきます。",
                "一枚一枚から作品の雰囲気が感じられます。",
                "写真を通じて色彩や空気感がダイレクトに伝わります。",
                "写真だけでも物語の余韻を感じ取ることができます。"
            ),
            choose(
                "構図や色彩が美しく、1枚1枚に作品の魅力が凝縮されています。",
                "光と影のコントラストが印象的です。",
                "美術や映像演出の細部までこだわりが感じられます。",
                "撮影の丁寧さが写真からも伝わってきます。"
            ),
            choose(
                "人物の表情や場面の切り取り方から感情の流れが伝わります。",
                "静止画でありながら物語の動きを感じさせます。",
                "数枚のスチールで物語の全体像が垣間見えます。",
                "登場人物の心情が写真からも伝わってきます。"
            ),
            choose(
                "以下のスチール写真から作品の魅力を一足先にご覧ください。",
                "本編を観る前に雰囲気を掴むのに役立ちます。",
                "スチールを見ながら鑑賞のポイントを探してみてください。",
                "気になるシーンを写真で先取りして楽しんでください。"
            )
        ]

    # ===============================
    # 評価と人気
    elif name == "評価と人気":
        base = [
            choose(
                f"映画『{title}』の評価は、観客の反応を端的に示しています。",
                "スコアから作品の支持度を推し量ることができます。",
                "点数は観客の期待と満足度を反映しています。",
                "数字を見ると作品への注目度が伝わってきます。"
            ),
            choose(
                "投票数が多いほど信頼性の高いデータになります。",
                "評価の平均値と参加者数を併せて見ると全体像が見えてきます。",
                "人気の度合いを知る上で投票数は重要な要素です。",
                "データの蓄積が作品の評価をより正確に示しています。"
            ),
            choose(
                "もちろん点数が全てではありません。最終的には自分の感性で判断するのが一番です。",
                "スコアは参考程度に、実際に観るのが一番の楽しみ方です。",
                "数字に惑わされず、自分の好みを信じることが大切です。",
                "数値よりも自分の直感を信じて観てみるのがおすすめです。"
            ),
            choose(
                "以下の評価データを参考にしてください。",
                "人気の傾向を数字で確認できます。",
                "データはあくまで目安としてご覧ください。",
                "気になる方は細かい統計もチェックしてみてください。"
            )
        ]

    # ===============================
    # ベストレビュー
    elif name == "ベストレビュー":
        base = [
            choose(
                "観客レビューには率直な感想が詰まっています。",
                "短いコメントからでも観客の感情が伝わります。",
                "レビューを読むことで作品の受け止められ方がよく分かります。",
                "リアルな感想から観客の熱量を感じられます。"
            ),
            choose(
                "好意的な意見もあれば、批判的な声もあります。それが映画の面白さです。",
                "賛否両論があるほど作品が話題になった証拠です。",
                "異なる視点が集まることで多面的な魅力が見えてきます。",
                "感想の多様性が作品の奥深さを示しています。"
            ),
            choose(
                "ここでは印象的なレビューをいくつかご紹介します。",
                "ネタバレを避けつつ感想をまとめました。",
                "短いレビューを中心に抜粋しています。",
                "観客の声をそのまま切り取ってみました。"
            ),
            choose(
                "読むうちに自分の好みや鑑賞ポイントが見えてくるはずです。",
                "他の人の視点を知ることで作品の理解が深まります。",
                "レビューから新しい発見が得られるかもしれません。",
                "違う視点に触れることで作品がより面白く感じられます。"
            )
        ]

    # ===============================
    # 予告編
    elif name == "予告編":
        base = [
            choose(
                "予告編は作品の雰囲気を一番わかりやすく伝えてくれます。",
                "短い時間に魅力が凝縮されています。",
                "予告を観るだけで作品のトーンが感じられます。",
                "映像の断片から作品の世界に引き込まれます。"
            ),
            choose(
                "ネタバレなしで雰囲気だけを味わえます。",
                "音楽と映像の組み合わせで作品の世界観が伝わります。",
                "リズムやテンポから本編の雰囲気を感じ取れます。",
                "短い時間でも物語のテーマをしっかりと感じられます。"
            ),
            choose(
                "短い中にもメッセージ性が込められています。",
                "数秒のシーンが作品のテーマを象徴することもあります。",
                "映像美やサウンドで作品の魅力が伝わります。",
                "一瞬の映像から深いテーマを感じられることもあります。"
            ),
            choose(
                "以下の予告編をご覧いただき、本編への期待を高めてください。",
                "映像をチェックして気になったらぜひ本編もご覧ください。",
                "予告編を観れば鑑賞する価値が一層感じられるはずです。",
                "少しでも気になったら本編もぜひチェックしてください。"
            )
        ]

    # ===============================
    # おすすめ映画
    elif name == "おすすめ映画":
        base = [
            choose(
                f"もし映画『{title}』を気に入ったなら、きっとこの作品も楽しめます。",
                f"映画『{title}』と似た雰囲気のおすすめ映画を集めました。",
                f"『{title}』と合わせて観るとさらに楽しめる作品をご紹介します。",
                f"『{title}』が好きな方にはぜひ観てほしい作品です。"
            ),
            choose(
                "今回はシンプルにタイトルとポスターのみを掲載しています。",
                "詳細な説明は省き、直感的に選べるようにしました。",
                "ポスターを見るだけで雰囲気を感じ取れるはずです。",
                "まずは見た目から直感で選んでみてください。"
            ),
            choose(
                "気になる作品があればチェックしてみてください。",
                "好みに合いそうな映画を見つけていただけると思います。",
                "視覚的に比較しながら楽しんでください。",
                "新しいお気に入りが見つかるかもしれません。"
            ),
            choose(
                "それではおすすめ映画をポスターとともにご覧ください。",
                "以下の作品群からお気に入りを探してみましょう。",
                "新しい映画との出会いがあるかもしれません。",
                "気軽に一覧からチェックしてみてください。"
            )
        ]

    # ===============================
    # その他 (デフォルト)
    else:
        base = [
            choose("重要なポイントを分かりやすく整理しました。"),
            choose("以下の内容を順番にチェックしていきましょう。"),
            choose("簡単に流れを追えるようにまとめています。"),
            choose("全体像をつかみやすいようにシンプルにまとめました。")
        ]

    return " ".join(base)


# ===============================
# 🎬 アウトロ生成 (6文ランダム, 日本語)
def make_outro_6(title, year, genres_str, director_names, keywords):
    director_one = director_names[0] if director_names else ""
    year_txt = year if year else "公開年不明"

    # 1. 締めの挨拶
    s1 = choose(
        f"ここまで <b>{title}</b>({year_txt}) についてご紹介しました。",
        f"今回は <b>{title}</b> の魅力を整理してみましたが、いかがでしたか？",
        f"<b>{title}</b>({year_txt}) の情報を一通りまとめました。",
        f"短くではありますが <b>{title}</b> の特徴をご紹介しました。",
        f"<b>{title}</b> に関する内容をお伝えしました。",
        f"<b>{title}</b> を振り返りながら楽しくまとめてみました。",
        f"映画 <b>{title}</b>({year_txt}) のお話はここまでとなります。",
        f"<b>{title}</b> の紹介はひとまずここで一区切りです。",
        f"最後まで <b>{title}</b> の世界を一緒に振り返れて嬉しかったです。",
        f"<b>{title}</b>({year_txt}) についてのまとめはここまでです。"
    )

    # 2. 内容まとめ
    s2 = choose(
        "あらすじから演出、キャストの演技まで一通り取り上げました。",
        "ストーリー、映像美、演技のバランスについて整理しました。",
        "全体の雰囲気やメッセージ性も簡単に触れました。",
        "作品の世界観やテーマ性についても少し紹介できました。",
        "細かな演出やキャラクターの魅力も振り返りました。",
        "音楽や映像効果の印象も一緒に見直しました。",
        "作品全体を通しての余韻についても触れてみました。",
        "キャラクター同士の関係性についても少しご紹介しました。"
    ) + " " + choose(
        f"{genres_str} ジャンルの魅力も確認できました。" if genres_str else "ジャンル的な魅力も感じられました。",
        f"{director_one} 監督ならではの演出も印象的でした。" if director_one else "演出面でも見どころが多かったです。",
        f"{director_one} 監督のタッチが作品全体に表れていました。" if director_one else "監督の工夫も感じられる場面が多かったです。",
        f"{director_one} 監督の独自の色が際立っていました。" if director_one else "演出スタイルの個性も伝わってきました。",
        f"{director_one} 監督の演出力を改めて感じることができました。" if director_one else "映像表現の工夫も随所に見られました。"
    )

    # 3. 評価について
    s3 = choose(
        "評価やスコアはあくまで参考にして、",
        "点数は一つの目安に過ぎないので、",
        "数字だけで判断せずに、",
        "ランキングに左右されすぎずに、",
        "世間の評価よりも自分の感覚を大切にして、",
        "レビューよりも自分の気持ちを信じて、",
        "周りの声に流されずに、",
        "数値やデータよりも体験そのものを重視して、"
    ) + choose(
        "最終的には自分の感性で楽しむのが一番です。",
        "好みに合わせて選ぶのが正解です。",
        "自分の直感を信じて鑑賞してみてください。",
        "心に響くかどうかが一番大切です。",
        "実際に観て感じることが一番の答えです。",
        "その時の気分や状況で楽しむのもおすすめです。",
        "自分の価値観に合うかどうかで判断してください。",
        "一度観てみることでしか分からない魅力があります。"
    )

    # 4. おすすめ映画
    s4 = choose(
        "似た雰囲気のおすすめ作品も紹介しましたので、ぜひ合わせて観てください。",
        "おすすめ映画もチェックして、映画体験を広げてみてください。",
        "気になる方は関連作品も探して観ていただけると楽しさが倍増します。",
        "他の作品と比べながら観ると、新しい発見があると思います。",
        "ぜひ次の映画選びの参考にしてみてください。",
        "ジャンルの近い作品を観比べるのも楽しいですよ。",
        "お気に入りの作品リストに追加してみてください。",
        "他の作品も合わせて観ることでより深い理解が得られます。"
    )

    # 5. キーワード
    s5 = choose(
        f"検索用キーワード: {', '.join(keywords[:8])}",
        f"関連ワード: {', '.join(keywords[:8])}",
        f"今回の主要キーワード: {', '.join(keywords[:8])}",
        f"覚えておきたいキーワード: {', '.join(keywords[:8])}",
        f"この記事で取り上げたキーワード: {', '.join(keywords[:8])}",
        f"参考になるキーワード: {', '.join(keywords[:8])}",
        f"メモしておきたいキーワード: {', '.join(keywords[:8])}",
        f"検索に便利なキーワード: {', '.join(keywords[:8])}"
    )

    # 6. 結び
    s6 = choose(
        "最後まで読んでいただきありがとうございました！🙂",
        "この記事が少しでも参考になれば嬉しいです。",
        "ぜひ感想をコメントで教えてください。",
        "次回も魅力的な映画を紹介しますのでお楽しみに！",
        "この記事が役に立ったと思ったらシェアしていただけると嬉しいです。",
        "ここまで読んでくださった方に心から感謝します。",
        "また別の作品も紹介していきますのでぜひチェックしてください。",
        "次の記事でもお会いできることを楽しみにしています。",
        "皆さんの映画ライフがより楽しくなることを願っています。",
        "読んでくださったあなたに感謝の気持ちを込めて🙂"
    )

    return " ".join([s1, s2, s3, s4, s5, s6])




import requests, random
import xml.etree.ElementTree as ET

def get_random_rss_links(rss_url="https://japan.appsos.kr/feeds/posts/default?alt=rss", count=5):
    try:
        r = requests.get(rss_url, timeout=10)
        r.raise_for_status()
        root = ET.fromstring(r.content)

        # RSS의 <item> 태그 추출
        items = root.findall(".//item")
        links = []
        for item in items:
            link = item.find("link").text if item.find("link") is not None else None
            title = item.find("title").text if item.find("title") is not None else "無題"
            if link:
                links.append((link, title))

        # 무작위로 count 개 선택
        selected = random.sample(links, min(count, len(links)))
        return selected

    except Exception as e:
        print("❌ RSS パース エラー:", e)
        return []

# RSS에서 무작위 5개 가져오기
rss_links = get_random_rss_links()

# HTML 생성
links_html = ""
for href, text in rss_links:
    links_html += f'<a href="{href}" style="color:#555555; font-weight:normal;">● {text}</a><br>\n'

related_block = f"""
<div style="background: rgb(239, 237, 233); border-radius: 8px;
            border: 2px dashed rgb(167, 162, 151);
            box-shadow: rgb(239, 237, 233) 0px 0px 0px 10px;
            color: #565656; font-weight: bold;
            margin: 2em 10px; padding: 2em;">
  <p style="border-bottom: 1px solid rgb(85, 85, 85); color: #555555;
            font-size: 16px; margin-bottom: 15px; padding-bottom: 5px;">
    ♡♥ 一緒に見たい記事
  </p>
  {links_html}
</div>
"""

# HTML ビルダー
def build_html(post, cast_count=10, stills_count=8):
    esc = html.escape
    title = esc(post.get("title") or post.get("original_title") or "タイトル不明")

    overview = esc(post.get("overview") or "あらすじ情報はまだ準備されていません。")
    release_date = esc(post.get("release_date") or "")
    year = release_date[:4] if release_date else ""
    runtime = post.get("runtime") or 0
    genres_list = [g.get("name","") for g in post.get("genres",[]) if g.get("name")]
    genres_str = "・".join(genres_list)
    tagline = esc(post.get("tagline") or "")
    adult_flag = bool(post.get("adult", False))

    # 制作国
    countries = [c.get("name","") for c in post.get("production_countries",[]) if c.get("name")]
    country_str = ", ".join(countries) if countries else "制作国情報なし"

    backdrop = img_url(post.get("backdrop_path"), "w1280")

    credits = post.get("credits", {}) or {}
    cast = credits.get("cast", [])[:cast_count]
    crew = credits.get("crew", [])
    directors = [c for c in crew if c.get("job") == "Director"]
    director_names = [esc(d.get("name","")) for d in directors]
    cast_names = [esc(p.get("name","")) for p in cast]

    backdrops = (post.get("images", {}) or {}).get("backdrops", [])
    backdrops = sorted(backdrops, key=lambda b: (b.get("vote_count",0), b.get("vote_average",0)), reverse=True)[:stills_count]

    cert = get_movie_release_cert(post["id"], bearer=BEARER, api_key=API_KEY)
    if not cert and adult_flag: cert = "R18"

    # キーワード生成
    base_keywords = []
    for w in (title.replace(":", " ").replace("-", " ").split()):
        if len(w) > 1:
            base_keywords.append(str(w))
    base_keywords += genres_list
    base_keywords += director_names[:2]
    base_keywords += cast_names[:3]
    if year: base_keywords.append(year)
    if cert: base_keywords.append(cert)
    base_keywords += ["レビュー", "評価", "出演者", "予告編", "スチール写真", "おすすめ映画"]

    # 重複削除
    seen = set()
    keywords = []
    for k in base_keywords:
        if isinstance(k, str) and k and k not in seen:
            keywords.append(k)
            seen.add(k)

    intro_6 = make_intro_6(title, year, genres_str, director_names, cast_names, cert, runtime, keywords)

    # 出演者テーブル
    cast_rows = []
    for p in cast:
        name = esc(p.get("name",""))
        ch = esc(p.get("character",""))
        prof = img_url(p.get("profile_path"), "w185")
        img_tag = f'<img src="{prof}" alt="{name}" style="width:72px;height:auto;border-radius:8px;">' if prof else ""
        cast_rows.append(
            f'<tr>'
            f'<td style="vertical-align:top;padding:8px 10px;white-space:nowrap;">{img_tag}</td>'
            f'<td style="vertical-align:top;padding:8px 10px;"><b>{name}</b><br><span style="color:#666;">{ch}</span></td>'
            f'</tr>'
        )
    cast_table = (
        '<table style="width:100%;border-collapse:collapse;border:1px solid #eee;">' +
        "".join(cast_rows or ['<tr><td style="padding:10px;">出演者情報がありません。</td></tr>']) +
        '</table>'
    )

    # スチール写真
    still_divs = []
    for b in backdrops:
        p = img_url(b.get("file_path"), "w780")
        if not p: continue
        still_divs.append(
            f'<div style="flex:0 0 49%;margin:0.5%;"><img src="{p}" alt="{title} スチール写真" style="width:100%;height:auto;border-radius:10px;"></div>'
        )
    stills_html = (
        '<div style="display:flex;flex-wrap:wrap;justify-content:space-between;">' +
        "".join(still_divs or ['<div style="padding:10px;">スチール写真がありません。</div>']) +
        '</div>'
    )

    # 評価と人気リード
    rating_lead = make_section_lead("評価と人気", title, year, genres_str, cert)

    vote_avg = post.get("vote_average", 0)
    vote_count = post.get("vote_count", 0)
    popularity = post.get("popularity", 0)

    rating_html = f"""
    <div style="background:linear-gradient(135deg,#f9f9f9,#ececec);
                border:2px solid #ddd;border-radius:15px;
                padding:30px;margin:20px 0;
                box-shadow:0 4px 12px rgba(0,0,0,0.08);
                text-align:center;">
    <div style="font-size:20px;font-weight:bold;margin-bottom:12px;color:#333;">
        ⭐ 評価 & 📊 人気スコア
    </div>
    <div style="font-size:18px;color:#222;margin:8px 0;">
        <b style="color:#ff9800;">平均評価:</b> {vote_avg:.1f}/10
    </div>
    <div style="font-size:16px;color:#555;margin:6px 0;">
        投票者数: {vote_count:,}人
    </div>
    <div style="font-size:18px;color:#0066cc;margin-top:10px;">
        <b>人気スコア:</b> {popularity:.1f}
    </div>
    </div>
    """

    # 予告編 (ja-JP → en-US fallback)
    # ✅ 안내文のバリエーション
    video_notice_variants = [
    "※ 自動取得のため、まれに予告編ではない動画が表示される場合があります。",
    "※ 下の動画は関連映像が混じることもありますのでご注意ください。",
    "※ まれに別の映像が再生されることがあります。ご了承ください。",
    "※ 予告編の読み込み中にエラーが発生すると、別の動画が表示される可能性があります。",
    "※ 正式な予告編と異なる動画が表示されることがあります。",
    "※ 関連する他の映像が一緒に出る場合があります。",
    "※ 自動検索のため、必ずしも公式予告編とは限りません。",
    "※ 稀に関連のない動画が表示されることがあります。",
    "※ 読み込みエラーにより正しく表示されない場合があります。",
    "※ 下の動画は参考用としてご覧ください。",
    "※ ネット環境によっては正しく再生されない場合があります。",
    "※ 掲載される動画は公開状況により変わることがあります。",
    "※ 動画の内容は予告なく変更される場合があります。",
    "※ 一部の映像は公式以外のものが含まれることがあります。",
    "※ ご利用の端末によっては表示に時間がかかる場合があります。",
    "※ 作品と直接関係のない動画が表示される可能性があります。",
    "※ 正確な情報は公式サイトや配給元の発表をご確認ください。",
    "※ こちらの動画はあくまで参考としてお楽しみください。",
    "※ 動画が再生できない場合は時間を置いて再試行してください。",
    "※ 自動的に取得された映像のため、内容に誤りがあることがあります。"
    ]


    # 🎬 予告編
        # 🎬 予告編
    video_html = ""
    video_lead = make_section_lead("予告編", title, year, genres_str, cert)

    # 1) TMDB公式 (Trailer/Official Trailer/Teaser 全部)
    videos = get_movie_videos_all(post["id"], bearer=BEARER, api_key=API_KEY)

    if videos:
        video_html += f"<p>{video_lead}</p>"
        for v in videos:
            yt_key = v.get("key")
            yt_name = html.escape(v.get("name") or "予告編")
            video_html += (
                f"<p><b>{yt_name}</b></p>"
                f"<iframe width='560' height='315' src='https://www.youtube.com/embed/{yt_key}' "
                f"frameborder='0' allowfullscreen></iframe><br>"
            )


    # 2) YouTube API検索 (最大2件, 보완용)

     
    yt_results = get_youtube_trailers(f"{title} 予告編", max_results=2)
    if yt_results:
        # 🔔 항상 안내문 출력 (YouTube 영상 나오기 전에)
        video_notice = random.choice(video_notice_variants)
        video_html += f"<br /><p style='color:#666;font-size:13px;'>{video_notice}</p><br />"

        for vid, vtitle in yt_results:
            video_html += (
                f"<p><b>{vtitle}</b></p>"
                f"<iframe width='560' height='315' src='https://www.youtube.com/embed/{vid}' "
                f"frameborder='0' allowfullscreen></iframe><br>"
            )





    # おすすめ映画
    recs = get_movie_recommendations(post["id"], lang=LANG, bearer=BEARER, api_key=API_KEY)
    rec_divs = []
    for r in recs[:6]:
        rtitle = esc(r.get("title") or r.get("original_title") or "")
        poster = img_url(r.get("poster_path"), "w185")
        link = f"https://japan.appsos.kr/search?q={urllib.parse.quote(rtitle)}"
        rec_divs.append(
            f'<div style="flex:0 0 32%;margin:1%;">'
            f'<a href="{link}" target="_blank">'
            f'<img src="{poster}" alt="{rtitle}" style="width:100%;border-radius:8px;"></a><br>'
            f'<a href="{link}" target="_blank" style="font-size:14px;color:#333;text-decoration:none;">{rtitle}</a>'
            f'</div>'
        )
    recs_html = (
    f'<h2>映画『{title}』のおすすめ作品</h2>'
    f'<p>{make_section_lead("おすすめ映画", title, year, genres_str, cert)}</p>'
    '<div style="display:flex;flex-wrap:wrap;justify-content:space-between;">'
    + "".join([
        f'<div style="flex:0 0 32%;margin-bottom:15px;text-align:center;">'
        f'<a href="https://japan.appsos.kr/search?q={urllib.parse.quote(r.get("title") or r.get("original_title") or "")}" target="_blank">'
        f'<img src="{img_url(r.get("poster_path"), "w185")}" '
        f'alt="{esc(r.get("title") or r.get("original_title") or "")}" '
        f'style="width:100%;border-radius:8px;"></a><br>'
        f'<a href="https://japan.appsos.kr/search?q={urllib.parse.quote(r.get("title") or r.get("original_title") or "")}" target="_blank" '
        f'style="font-size:14px;color:#333;text-decoration:none;">'
        f'{esc(r.get("title") or r.get("original_title") or "")}</a>'
        f'</div>'
        for r in recs[:6]
    ]) if recs else "<p>おすすめ映画情報はありません。</p>"
    + "</div>"
    )





    # アウトロ
    outro_6 = make_outro_6(title, year, genres_str, director_names, keywords)

    # 最終 HTML
    blog_title1 = f"映画 {title} ({year}) あらすじ 出演者 主人公 予告編"
    hashtags = make_hashtags_from_title(blog_title1)

    html_out = f"""
<p>{intro_6}</p>
<!--more--><br />
{"<p><img src='"+backdrop+"' style='width:100%;height:auto;border-radius:12px;'></p>" if backdrop else ""}
{"<p><i>"+html.escape(tagline)+"</i></p>" if tagline else ""}

<br /><br /><br />
<h2>映画『{title}』のあらすじ</h2>
<p><b>制作国:</b> {country_str} | <b>ジャンル:</b> {genres_str if genres_str else "ジャンル情報なし"}</p>
<p>{make_section_lead("あらすじ", title, year, genres_str, cert)}</p>

{f'''<div class="ottistMultiRelated">
  <a class="extL alt" href="https://japan.appsos.kr/search/label/{year}?&max-results=10">
    <span style="font-size: medium;"><strong>{year}年おすすめ映画を見に行く</strong></span>
    <i class="fas fa-link 2xs"></i>
  </a>
</div>''' if year else ''}
<div style="background:#fafafa;border:2px solid #ddd;border-radius:12px;
            padding:10px 18px 25px;margin:18px 0;line-height:1.7;color:#333;
            box-shadow:0 3px 8px rgba(0,0,0,0.05);">
  <p style="font-weight:bold;font-size:16px;margin-bottom:10px;">🎬 {title} あらすじ</p>
  {overview}
</div>
<br />
{hashtags}

<br /><br /><br />
<h2>映画『{title}』の出演者</h2>
<p>{make_section_lead("出演者", title, year, genres_str, cert, extras={"cast_top": cast_names})}</p>
{cast_table}
<br />
{hashtags}

<br /><br /><br />
<h2>映画『{title}』のスチール写真</h2>
<p>{make_section_lead("スチール写真", title, year, genres_str, cert)}</p>
{f'''<div class="ottistMultiRelated">
  <a class="extL alt" href="https://japan.appsos.kr/search/label/{urllib.parse.quote(genres_list[0])}?&max-results=10">
    <span style="font-size: medium;"><strong>おすすめ {genres_list[0]} 映画を見に行く</strong></span>
    <i class="fas fa-link 2xs"></i>
  </a>
</div>''' if genres_list else ''}
{stills_html}
<br />
{hashtags}

<br /><br /><br />
<h2>映画『{title}』の評価と予告編</h2>
<p>{rating_lead}</p>
{rating_html}
{video_html}

<br /><br /><br />
{recs_html}

<br />
<p>{outro_6}</p>
{related_block}


<p style="font-size:12px;color:#666;">
本コンテンツは <a href="https://www.themoviedb.org/" target="_blank" style="color:#666;text-decoration:underline;">TMDB</a> のデータに基づいて作成されています。
</p>
"""

  

    

    return textwrap.dedent(html_out).strip()



# ===============================
# Blogger 認証/投稿
from googleapiclient.discovery import build
import google.oauth2.credentials

CLIENT_SECRET_FILE = r"cc.json"   # 기존 cc.json 유지 (로컬 refresh_token 발급용)
BLOGGER_TOKEN_JSON = "blogger_token.json"  # 새 JSON 토큰 파일
SCOPES = ["https://www.googleapis.com/auth/blogger"]

# Blogger 인증 (JSON 토큰 방식)
def get_blogger_service():
    try:
        if not os.path.exists(BLOGGER_TOKEN_JSON):
            raise FileNotFoundError("❌ blogger_token.json 파일이 없습니다. 먼저 로컬에서 발급해 업로드하세요.")

        with open(BLOGGER_TOKEN_JSON, "r", encoding="utf-8") as f:
            token_data = json.load(f)

        creds = google.oauth2.credentials.Credentials.from_authorized_user_info(token_data, SCOPES)
        return build("blogger", "v3", credentials=creds)
    except Exception as e:
        print(f"❌ Blogger 인증 실패: {e}", file=sys.stderr)
        raise


# ===============================
# Excel ヘルパー（G列に '완'）
DONE_COL = 7         # G列
DONE_MARK = "완"     # 표시는 '완' (한국어)


def post_to_blogger(service, blog_id, title, html_content, labels=None, is_draft=False):
    """
    Blogger API로 글 발행
    """
    body = {
        "kind": "blogger#post",
        "title": title,
        "content": html_content,
    }
    if labels:
        body["labels"] = labels

    try:
        post = service.posts().insert(
            blogId=blog_id,
            body=body,
            isDraft=is_draft
        ).execute()
        return post
    except Exception as e:
        print(f"❌ Blogger API 포스트 실패: {e}", file=sys.stderr)
        raise


def find_next_row(ws):
    """
    G열이 '완'이 아닌 첫 번째 데이터 행을 찾는다 (2행부터).
    반환: (row_index, movie_id:int)
    """
    for row_idx in range(2, ws.max_row + 1):
        done_val = str(ws.cell(row=row_idx, column=DONE_COL).value or "").strip()  # G열
        movie_raw = ws.cell(row=row_idx, column=2).value                           # B열 (MOVIE_ID)
        if done_val == DONE_MARK:
            continue
        if movie_raw is None or str(movie_raw).strip() == "":
            continue
        try:
            movie_id = int(str(movie_raw).strip())
        except:
            continue
        return row_idx, movie_id
    return None, None

def mark_done(ws, row_idx):
    ws.cell(row=row_idx, column=DONE_COL, value=DONE_MARK)  # G열 = '완'

# ===============================
# メイン実行部
def main_once():
    random.seed()

    # 1) 엑셀에서 대상 행 찾기
    if not os.path.exists(EXCEL_PATH):
        print(f"엑셀 파일을 찾지 못했습니다: {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    target_row, movie_id = find_next_row(ws)
    if not movie_id:
        print("처리할 행이 없습니다. (모든 행이 '완')")
        return False  # 더 이상 처리할 게 없으니 종료

    print(f"👉 대상 행: {target_row} (MOVIE_ID={movie_id})")

    # 2) TMDB 상세 수집
    try:
        post = get_movie_bundle(movie_id, lang=LANG, bearer=BEARER, api_key=API_KEY)
    except Exception as e:
        print(f"TMDB 요청 실패: {e}", file=sys.stderr)
        return True  # 다음 턴 계속

    # 3) HTML 생성
    try:
        html_out = build_html(post, cast_count=CAST_COUNT, stills_count=STILLS_COUNT)
    except Exception as e:
        print(f"HTML 생성 오류: {e}", file=sys.stderr)
        return True

    # 4) 블로그 제목(일본어)
    title = (post.get("title") or post.get("original_title") or f"movie_{movie_id}")
    year = (post.get("release_date") or "")[:4]
    blog_title = f"映画 {title} ({year}) あらすじ 出演者 主人公 予告編"

    # 5) Blogger 발행
    post_success = False
    try:
        service = get_blogger_service()
        genres_list = [g.get("name","") for g in post.get("genres",[]) if g.get("name")]
        year = (post.get("release_date") or "")[:4]

        labels = ["映画"]
        if year:
            labels.append(year)
        if genres_list:
            labels.extend(genres_list)

        res = post_to_blogger(service, BLOG_ID, blog_title, html_out, labels=labels, is_draft=False)
        post_url = res.get("url", "(URL 미확인)")
        print(f"✅ 발행 완료: {post_url}")
        post_success = True
    except Exception as e:
        print(f"Blogger 발행 실패: {e}", file=sys.stderr)
       

         # 👉 실패 시 HTML 로컬 저장 (D:\py\image_wd 고정)
        safe_title = "".join(c if c.isalnum() else "_" for c in blog_title)[:50]
        save_dir = r"D:\py\image_wd"
        os.makedirs(save_dir, exist_ok=True)  # 폴더 없으면 자동 생성
        filename = f"failed_post_{movie_id}_{safe_title}.html"
        filepath = os.path.join(save_dir, filename)


        try:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(html_out)
            print(f"💾 발행 실패 → HTML 로컬 저장 완료\n    저장 위치: {filepath}")
        except Exception as fe:
            print(f"❌ HTML 저장 실패: {fe}", file=sys.stderr)

    # 6) 엑셀 G열 '완' 표시 후 저장 (✅ 성공시에만)
    if post_success:
        try:
            mark_done(ws, target_row)
            wb.save(EXCEL_PATH)
            print(f"✅ 엑셀 저장 완료: {EXCEL_PATH} (행 {target_row} G열='{DONE_MARK}')")
        except Exception as e:
            print(f"엑셀 저장 실패: {e}", file=sys.stderr)

    return True  # 다음 루프 진행



if __name__ == "__main__":
    for i in range(POST_COUNT):
        print(f"\n🚀 {i+1}/{POST_COUNT} 번째 포스팅 시작")
        ok = main_once()
        if not ok:
            print("📌 더 이상 처리할 데이터가 없어 종료합니다.")
            break

        if i < POST_COUNT - 1 and POST_DELAY_MIN > 0:
            print(f"⏳ {POST_DELAY_MIN}분 대기 후 다음 포스팅...")
            time.sleep(POST_DELAY_MIN * 60)




